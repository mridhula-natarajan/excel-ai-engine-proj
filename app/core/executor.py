import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, Any, Optional
from app.core.logger import get_logger
from app.core.executor_helpers import (
    _do_aggregate, _do_math, _do_filter, _do_pivot,
    _do_unpivot, _do_join, _do_date_ops, _do_text_analysis,
    to_serializable, derive_missing_columns_with_llm
)

logger = get_logger(__name__)

def execute_plan(df: pd.DataFrame, plan: Dict[str, Any], other_tables: Optional[Dict[str, pd.DataFrame]] = None) -> Dict[str, Any]:
    """Executes the predicted operation plan on the given Excel dataframe"""
    logger.info(" Executing plan: %s", plan)
    df, derivation_report = derive_missing_columns_with_llm(plan, df, logger)
    if derivation_report:
        logger.info("Derivation report: %s", derivation_report)

    try:
        op = plan.get("operation", "").strip().lower()
        params = plan.get("parameters", {})
        logger.info("Detected operation: %s", op)

        # Dispatch to appropriate executor
        if op == "aggregate":
            result = _do_aggregate(df, params)
        elif op == "math":
            result = _do_math(df, params)
        elif op == "filter":
            result = _do_filter(df, params)
        elif op == "pivot":
            result = _do_pivot(df, params)
        elif op == "unpivot":
            result = _do_unpivot(df, params)
        elif op == "join":
            result = _do_join(df, params, other_tables)
        elif op == "date_ops":
            result = _do_date_ops(df, params)
        elif op == "text_analysis":
            result = _do_text_analysis(df, params)
        elif op in ["describe", "sample"]:
            result = {
                "status": "ok",
                "result_df": df,
                "preview": to_serializable(df, max_rows=10),
                "message": "describe/sample"
            }
        elif op == "multi_step":
            current_df = df.copy()
            for step in plan.get("parameters", {}).get("steps", []):
                logger.info(" Executing sub-step: %s", step)
                out = execute_plan(current_df, step, other_tables=other_tables)
                if out["status"] != "ok":
                    return out
                if out.get("result_df") is not None:
                    current_df = out["result_df"]
            result = {
                "status": "ok",
                "result_df": current_df,
                "preview": to_serializable(current_df, max_rows=20),
                "message": "multi_step executed"
            }
        else:
            logger.warning("Unsupported operation: %s", op)
            return {"status": "error", "message": f"Unsupported operation: {op}"}

        logger.info(" Operation %s executed successfully.", op)

        # Saving the output in excel file along with input data
        input_path = plan.get("input_path")
        result_df = result.get("result_df")

        if result.get("status") == "ok" and "result_df" in result:
            try:
                os.makedirs("results", exist_ok=True)
                input_path = plan.get("input_path")

                # handle case where no input file is provided 
                if input_path is None:
                    logger.warning("Skipping Excel save — no input_path provided (e.g., multi-step or ad-hoc query).")
                    result["message"] = "Result returned (no Excel file created)"
                    result["preview"] = to_serializable(result.get("result_df"), max_rows=10)
                    return {
                        "status": "ok",
                        "result_df": result["result_df"],
                        "preview": result["preview"],
                        "message": result["message"]
                    }

                input_name = os.path.basename(input_path)
                result_path = os.path.join("results", f"result_{input_name}")

                # Make a copy of the input Excel 
                shutil.copy2(input_path, result_path)

                # find the target sheet
                wb = load_workbook(result_path)
                sheet_name = wb.sheetnames[0]  # assuming first sheet
                ws = wb[sheet_name]

                result_df = result["result_df"]
                same_shape = len(result_df) == len(df)

                if same_shape:
                    existing_cols = len(ws[1])
                    for j, col in enumerate(result_df.columns, start=existing_cols + 1):
                        ws.cell(row=1, column=j, value=col)
                        for i, val in enumerate(result_df[col].tolist(), start=2):
                            ws.cell(row=i, column=j, value=val)
                else:
                    logger.info(" Writing summary report")
                    summary_sheet = wb["Summary Report"] if "Summary Report" in wb.sheetnames else wb.create_sheet("Summary Report")
                    summary_sheet.append(["Summary Report"])
                    # summary_sheet.append([])

                    try:
                        preview = result.get("preview", [])
                        if preview and isinstance(preview[0], dict):
                            for row in preview:
                                line = ", ".join(f"{k}: {v}" for k, v in row.items())
                                summary_sheet.append([line])
                        else:
                            summary_sheet.append([str(preview)])
                    except Exception as e:
                        summary_sheet.append([f"(Could not format summary: {e})"])

                wb.save(result_path)
                wb.close()

                result["file_path"] = result_path
                result["result_file"] = os.path.basename(result_path)
                result["message"] = f"Result written to {result_path}"
                logger.info(f" Result written to {result_path}")

            except Exception as e:
                logger.error(" Failed to save result: %s", e)
                result["message"] = f"Save failed: {e}"
                
        return result

    except Exception as e:
        logger.exception(" Exception during plan execution: %s", e)
        return {"status": "error", "message": str(e)}
